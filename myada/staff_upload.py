import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import StudentResult, Course, Student
from .forms import FileUploadForm
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook

def upload_excel(request):
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file_uploaded = form.cleaned_data['file']
            workbook = load_workbook(file_uploaded)

            try:
                # Read the Excel file into a pandas DataFrame
                #df = pd.read_excel(excel_file)

                # Iterate through the DataFrame rows and save data to StudentResult
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for index, row in enumerate(sheet.iter_rows(values_only=True)):
                        if index == 0:
                            continue
                        student_id = row[0]  
                        mat_number = row[1]
                        course_name = row[2]  
                        test = row[3]  
                        exam = row[4]  
                        try:
                            student_instance = Student.objects.get(id=student_id)
                        except Student.DoesNotExist:
                            messages.error(request, f'Student with name {student_id} does not exist.')
                            continue  # Skip to the next iteration

                        # Get the Subject instance based on the subject name
                        course_instance = Course.objects.get(name=course_name)

                        # Create or update StudentResult entry
                        student_result, created = StudentResult.objects.update_or_create(
                            student=student_instance,
                            course=course_instance,
                            defaults={'test': test, 'exam': exam, 'mat_number': mat_number}
                        )

                messages.success(request, 'Excel file uploaded successfully.')
                #return redirect('your_redirect_url')  # Replace 'your_redirect_url' with the actual URL you want to redirect to
            except Exception as e:
                messages.error(request, f'Error processing Excel file: {e}')
        else:
            messages.error(request, 'Invalid form submission. Please check the uploaded file.')

    else:
        form = FileUploadForm()

    return render(request, 'staff_template/result_upload.html', {'form': form})
